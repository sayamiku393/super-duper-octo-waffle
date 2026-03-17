import psycopg2
from sqlalchemy import text, MetaData, create_engine, insert
from openpyxl import load_workbook
from datetime import datetime

engine = create_engine("postgresql+psycopg2://postgres:umiko@localhost:5432/try_4")
meta = MetaData()
meta.reflect(bind=engine)

orders = meta.tables["orderz"]
users = meta.tables["userz"]
relations = meta.tables["relationz"]
points = meta.tables["pointz"]
items = meta.tables["itemz"]


def check__(param):
    s = str(param).split()[0]
    for fmt in ("%d-%m-%Y", "%Y-%m-%d"):
        try:
            result = datetime.strptime(s, fmt).date()
            return result
        except:
            continue

def call__(param):
    with engine.begin() as conn:
        result = conn.execute(text(f"select id from userz where fio = \'{param}\'")).fetchone()
        return result[0]

if __name__ == '__main__':
    with engine.begin() as conn:
        conn.execute(text("truncate table userz, pointz, itemz, orderz, relationz"))

    ord_doc = load_workbook(r"C:\Users\Neco-Tyan\Desktop\import\Заказ_import.xlsx").active
    usr_doc = load_workbook(r"C:\Users\Neco-Tyan\Desktop\import\user_import.xlsx").active
    pts_doc = load_workbook(r"C:\Users\Neco-Tyan\Desktop\import\Пункты выдачи_import.xlsx").active
    itm_doc = load_workbook(r"C:\Users\Neco-Tyan\Desktop\import\Tovar.xlsx").active

    ord_arr = []
    usr_arr = []
    pts_arr = []
    itm_arr = []
    rel_arr = []

    for i, row in enumerate(pts_doc.iter_rows(min_row=2, values_only=True)):
        if row[0] is None:
            break
        pts_arr.append({
            "id": i,
            "address": row[0]
        })
    with engine.begin() as conn:
        conn.execute(insert(points), pts_arr)

    for i, row in enumerate(usr_doc.iter_rows(min_row=2, values_only=True)):
        if row[0] is None:
            break
        usr_arr.append({
            "id": i,
            "role": row[0],
            "fio": row[1],
            "login": row[2],
            "password": row[3]
        })
    with engine.begin() as conn:
        conn.execute(insert(users), usr_arr)

    for row in itm_doc.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        itm_arr.append({
            "art": row[0],
            "model": row[1],
            "measure": row[2],
            "price": row[3],
            "supplier": row[4],
            "manufactr": row[5],
            "category": row[6],
            "discount": row[7],
            "amount": row[8],
            "descr": row[9],
            "photopath": f"C:\\Users\\Neco-Tyan\\Desktop\\import\\pics\\{row[10]}" if row[10] is not None else f"C:\\Users\\Neco-Tyan\\Desktop\\import\\pics\\picture.png"
        })
    with engine.begin() as conn:
        conn.execute(insert(items), itm_arr)

    for row in ord_doc.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        try:
            ord_ = check__(row[2])
            del_ = check__(row[3])
            fio = call__(row[5])
        except:
            continue
        ord_arr.append({
            "id": row[0],
            "ord_date": ord_,
            "del_date": del_,
            "point_id": row[4],
            "client_id": fio,
            "code": row[6],
            "status": row[7]
        })

        raw = str(row[1]).split(", ")
        for i in range (0, len(raw) - 1, 2):
            rel_arr.append({
                "order_id": row[0],
                "item_art": raw[i],
                "amt": raw[i + 1]
            })
    with engine.begin() as conn:
        conn.execute(insert(orders), ord_arr)
        conn.execute(insert(relations), rel_arr)